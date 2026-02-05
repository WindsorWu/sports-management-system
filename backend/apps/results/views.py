"""
成绩应用视图

提供成绩管理的完整REST API接口，包括成绩录入、查询、导出、批量导入等功能
"""
import re
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.db.models import Case, When, IntegerField
from rest_framework.exceptions import PermissionDenied
from openpyxl import load_workbook
from django.db import transaction

from .models import Result
from .serializers import ResultSerializer, ResultCreateSerializer, ResultListSerializer
from utils.permissions import IsAdmin, IsAdminOrReferee
from utils.export import export_results
from apps.events.models import RefereeEventAccess, Event
from apps.registrations.models import Registration


HEADER_ALIASES = {
    'event': {alias.lower() for alias in {
        '赛事名称', '赛事', '比赛名称', 'event name', 'event', 'match name'
    }},
    'participant': {alias.lower() for alias in {
        '参赛者', '选手', '运动员', 'participant', 'athlete', '选手姓名', 'name'
    }},
    'round': {alias.lower() for alias in {
        '轮次', 'round', 'round type', '阶段', '赛次'
    }},
    'score': {alias.lower() for alias in {
        '成绩', 'score', 'result'
    }},
    'rank': {alias.lower() for alias in {
        '排名', 'rank', 'position'
    }}
}

ROUND_TYPE_KEYWORDS = [
    ('semifinal', ['半决赛', 'semifinal']),
    ('final', ['决赛', 'final']),
    ('preliminary', ['预赛', '初赛', 'preliminary'])
]


def build_column_mapping(header_row):
    """
    构建Excel列映射关系
    
    将Excel表头与系统字段建立映射关系，支持多种表头名称
    
    参数:
        header_row: Excel第一行的表头数据
        
    返回:
        mapping: 字段名到列索引的映射字典
        missing: 缺失的必填字段列表
    """
    mapping = {}
    for idx, header in enumerate(header_row or []):
        normalized = (header or '').strip().lower()
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapping[field] = idx
                break
    missing = [field for field in HEADER_ALIASES if field not in mapping]
    return mapping, missing


def safe_str(value):
    """安全地将值转换为字符串，处理None和空值"""
    if value is None:
        return ''
    return str(value).strip()


def normalize_round_type(value):
    """
    规范化轮次类型
    
    将Excel中的轮次描述转换为系统标准格式
    支持中英文关键词匹配
    """
    if value is None:
        return None
    text = str(value).strip().lower()
    for target, keywords in ROUND_TYPE_KEYWORDS:
        for keyword in keywords:
            if keyword in text:
                return target
    return None


def build_candidate_names(raw_value):
    """
    构建参赛者姓名候选列表
    
    从Excel单元格提取可能的姓名，包括括号中的别名
    例如: "张三(zhangsan)" -> ["张三(zhangsan)", "zhangsan"]
    """
    if raw_value in (None, ''):
        return []
    text = str(raw_value).strip()
    if not text:
        return []
    names = [text]
    for match in re.findall(r"\(([^)]+)\)", text):
        candidate = match.strip()
        if candidate and candidate not in names:
            names.append(candidate)
    return names


def _match_registration(queryset, field, candidate):
    if not candidate:
        return None, None
    matches = queryset.filter(**{f"{field}__iexact": candidate})
    count = matches.count()
    if count == 1:
        return matches.first(), None
    if count > 1:
        return None, f'参赛者"{candidate}"匹配到多条报名记录，请确保唯一'
    return None, None


def find_registration_for_participant(event, candidates):
    """
    根据候选姓名查找报名记录
    
    在赛事的已审核报名中查找匹配的报名记录
    支持按参赛者姓名、真实姓名、用户名等多种方式匹配
    
    参数:
        event: 赛事对象
        candidates: 候选姓名列表
        
    返回:
        registration: 匹配的报名记录，未找到返回None
        error: 错误信息，无错误返回None
    """
    approved = Registration.objects.filter(event=event, status='approved').select_related('user')
    for candidate in candidates:
        for field in ['participant_name', 'user__real_name', 'user__username']:
            registration, error = _match_registration(approved, field, candidate)
            if error:
                return None, error
            if registration:
                return registration, None
    return None, '找不到与参赛者匹配的报名记录，请确认姓名或用户名'


class ResultViewSet(viewsets.ModelViewSet):
    """
    成绩视图集
    
    提供成绩管理的完整CRUD操作和扩展功能
    
    主要功能:
        - 成绩录入: 裁判录入运动员成绩
        - 成绩查询: 支持多条件筛选和搜索
        - 成绩公开: 控制成绩的公开状态
        - 成绩导出: 导出Excel格式的成绩表
        - 批量导入: 通过Excel批量导入成绩
        - 排行榜: 获取赛事排行榜
        
    权限控制:
        - 列表/详情: 任何人可访问（但只能看已公开的）
        - 创建/更新/删除: 需要管理员或裁判权限
        - 公开/导出/批量操作: 需要管理员或裁判权限
        - 裁判只能操作被分配的赛事
        
    使用场景:
        - 裁判录入比赛成绩
        - 生成赛事排行榜
        - 运动员查看个人成绩
        - 批量导入成绩数据
    """
    queryset = Result.objects.select_related('event', 'user', 'registration', 'recorded_by').all()
    serializer_class = ResultSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['event', 'user', 'round_type', 'is_published']
    search_fields = [
        'user__username', 'user__real_name', 'event__title',
        'score', 'award'
    ]
    ordering_fields = ['created_at', 'rank', 'score']
    ordering = ['event', 'rank']

    def get_permissions(self):
        """
        根据不同操作动态设置权限
        
        权限说明:
            - list/retrieve: 任何人可访问，但普通用户只能看已公开的
            - create/update/destroy: 需要管理员或裁判权限
            - publish/export/bulk_*: 需要管理员或裁判权限
            - 裁判受到赛事分配限制
        """
        if self.action in ['list', 'retrieve']:
            # 列表和详情允许任何人访问（但只能看到已公开的）
            permission_classes = [AllowAny]
        elif self.action in ['create', 'update', 'partial_update', 'destroy', 'publish', 'export', 'pending_results_count', 'import_results', 'bulk_publish', 'bulk_delete']:
            # 创建、更新、删除、公开、导出需要管理员或裁判权限
            permission_classes = [IsAdminOrReferee]
        else:
            # 其他操作需要认证
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def get_serializer_class(self):
        """根据不同操作返回不同的序列化器"""
        if self.action == 'create':
            return ResultCreateSerializer
        elif self.action == 'list':
            return ResultListSerializer
        return ResultSerializer

    def get_referee_event_ids(self, user):
        """
        获取裁判被分配的赛事ID列表
        
        如果用户不是裁判，返回None（不受限制）
        """
        if not (user and user.is_authenticated and user.user_type == 'referee'):
            return None
        return list(RefereeEventAccess.objects.filter(referee=user).values_list('event_id', flat=True))

    def apply_referee_filter(self, queryset, user):
        """
        应用裁判权限过滤
        
        限制裁判只能操作被分配的赛事
        """
        referee_ids = self.get_referee_event_ids(user)
        if referee_ids is None:
            return queryset
        return queryset.filter(event_id__in=referee_ids)

    def get_queryset(self):
        """
        限制普通用户只能看到已公开的成绩
        
        权限规则:
            - 管理员/组织者: 可以看到所有成绩
            - 裁判: 只能看到被分配赛事的成绩
            - 普通用户: 只能看到已公开的成绩
        """
        user = self.request.user
        if user.is_authenticated and (user.is_superuser or user.user_type in ['admin', 'organizer']):
            return self.queryset
        base = self.queryset.filter(is_published=True)
        return self.apply_referee_filter(base, user)

    def create(self, request, *args, **kwargs):
        """
        创建成绩记录
        
        业务逻辑:
            1. 验证裁判是否有权限录入该赛事成绩
            2. 验证报名记录是否有效
            3. 自动设置录入人为当前用户
            4. 自动从报名记录获取用户信息
        """
        referee_events = self.get_referee_event_ids(request.user)
        event_id = request.data.get('event')
        if referee_events is not None and event_id and int(event_id) not in referee_events:
            raise PermissionDenied('该裁判未被分配该赛事')
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = serializer.save()
        return Response({
            'message': '成绩录入成功',
            'result': ResultSerializer(result).data
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['put'])
    def publish(self, request, pk=None):
        """
        公开成绩
        PUT /api/results/{id}/publish/
        """
        result = self.get_object()
        result.is_published = True
        result.save()

        return Response({
            'message': '成绩已公开',
            'result': ResultSerializer(result).data
        })

    @action(detail=True, methods=['put'])
    def unpublish(self, request, pk=None):
        """
        取消公开成绩
        PUT /api/results/{id}/unpublish/
        """
        result = self.get_object()
        result.is_published = False
        result.save()

        return Response({
            'message': '成绩已取消公开',
            'result': ResultSerializer(result).data
        })

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        导出成绩表
        GET /api/results/export/?event={event_id}
        """
        # 获取查询参数
        event_id = request.query_params.get('event')
        round_type = request.query_params.get('round_type')

        queryset = Result.objects.select_related('event', 'user', 'registration').all()
        queryset = self.apply_referee_filter(queryset, request.user)

        if event_id:
            queryset = queryset.filter(event_id=event_id)
        if round_type:
            queryset = queryset.filter(round_type=round_type)

        round_order = Case(
            When(round_type='final', then=1),
            When(round_type='semifinal', then=2),
            When(round_type='preliminary', then=3),
            default=4,
            output_field=IntegerField()
        )
        queryset = queryset.order_by(round_order, 'rank', 'score')

        # 使用导出工具导出
        return export_results(queryset)

    @action(detail=False, methods=['get'])
    def leaderboard(self, request):
        """
        获取排行榜
        
        GET /api/results/leaderboard/?event={event_id}&round_type={round_type}
        
        功能说明:
            - 获取指定赛事和轮次的前10名成绩
            - 按排名升序排列
            - 只返回已公开的成绩
            
        参数:
            - event: 赛事ID（必填）
            - round_type: 轮次类型（可选，默认为final）
            
        返回:
            前10名的成绩列表
        """
        event_id = request.query_params.get('event')
        round_type = request.query_params.get('round_type', 'final')

        if not event_id:
            return Response({
                'error': '请提供赛事ID'
            }, status=status.HTTP_400_BAD_REQUEST)

        results = self.apply_referee_filter(self.queryset, request.user).filter(
            event_id=event_id,
            round_type=round_type,
            is_published=True
        ).order_by('rank')[:10]  # 只返回前10名

        serializer = ResultListSerializer(results, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_results(self, request):
        """
        获取当前用户的成绩
        
        GET /api/results/my_results/
        
        功能说明:
            - 获取当前登录用户的所有成绩
            - 只返回已公开的成绩
            - 按创建时间倒序排列
            
        返回:
            用户的成绩列表
        """
        if not request.user.is_authenticated:
            return Response({
                'error': '请先登录'
            }, status=status.HTTP_401_UNAUTHORIZED)

        base_results = Result.objects.filter(
            user=request.user,
            is_published=True
        ).select_related('event', 'registration')
        results = self.apply_referee_filter(base_results, request.user)

        serializer = ResultSerializer(results, many=True)
        return Response(serializer.data)

    def get_pending_results_count(self, event_ids=None):
        """
        获取待录入成绩的运动员数量
        
        计算已审核通过但尚未录入成绩的报名数量
        
        参数:
            event_ids: 赛事ID列表，用于限制统计范围
            
        返回:
            待录入成绩的数量
        """
        queryset = Result.objects.select_related('event', 'registration')
        if event_ids:
            queryset = queryset.filter(event_id__in=event_ids)
        recorded_ids = set(queryset.values_list('registration_id', flat=True))
        reg_queryset = Registration.objects.filter(status='approved')
        if event_ids:
            reg_queryset = reg_queryset.filter(event_id__in=event_ids)
        if recorded_ids:
            reg_queryset = reg_queryset.exclude(id__in=recorded_ids)
        return reg_queryset.count()

    @action(detail=False, methods=['get'])
    def pending_results_count(self, request):
        """
        获取待录入成绩数量（仅管理员/裁判）
        
        GET /api/results/pending_results_count/
        
        功能说明:
            - 管理员可以看到所有待录入数量
            - 裁判只能看到自己负责赛事的待录入数量
            - 用于提醒和统计
            
        返回:
            待录入成绩的数量
        """
        user = request.user
        if user.is_authenticated and (user.is_superuser or user.user_type in ['admin', 'organizer']):
            count = self.get_pending_results_count()
            return Response({'count': count})

        event_ids = self.get_referee_event_ids(user)
        if not event_ids:
            return Response({'count': 0})
        count = self.get_pending_results_count(event_ids)
        return Response({'count': count})

    @action(detail=False, methods=['post'])
    def bulk_publish(self, request):
        """
        批量公开成绩
        
        POST /api/results/bulk_publish/
        Body: {ids: [1, 2, 3]}
        
        功能说明:
            - 批量设置成绩为公开状态
            - 公开后运动员可以看到成绩
            - 裁判受到赛事分配限制
            
        参数:
            - ids: 要公开的成绩ID列表
            
        返回:
            成功公开的数量
        """
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'error': '请提供要公开的成绩ID列表'}, status=status.HTTP_400_BAD_REQUEST)
        queryset = Result.objects.filter(id__in=ids)
        queryset = self.apply_referee_filter(queryset, request.user)
        if not queryset.exists():
            return Response({'error': '没有可以公开的数据'}, status=status.HTTP_400_BAD_REQUEST)
        updated = queryset.update(is_published=True)
        return Response({'message': f'成功公开{updated}条成绩', 'updated': updated})

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """
        批量删除成绩
        
        POST /api/results/bulk_delete/
        Body: {ids: [1, 2, 3]}
        
        功能说明:
            - 批量删除多条成绩记录
            - 使用数据库事务确保数据一致性
            - 裁判受到赛事分配限制
            
        参数:
            - ids: 要删除的成绩ID列表
            
        返回:
            成功删除的数量
            
        注意事项:
            - 删除操作不可恢复
        """
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'error': '请提供要删除的成绩ID列表'}, status=status.HTTP_400_BAD_REQUEST)
        queryset = Result.objects.filter(id__in=ids)
        queryset = self.apply_referee_filter(queryset, request.user)
        if not queryset.exists():
            return Response({'error': '没有可以删除的数据'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            deleted, _ = queryset.delete()
        return Response({'message': f'成功删除{deleted}条成绩', 'deleted': deleted})

    @action(detail=False, methods=['post'], url_path='import')
    def import_results(self, request):
        """
        通过Excel批量导入成绩
        
        POST /api/results/import/
        Form-Data: file=成绩表.xlsx, context_event=赛事ID
        
        功能说明:
            - 支持从Excel文件批量导入成绩
            - 自动匹配参赛者和报名记录
            - 支持多种表头名称（中英文）
            - 自动识别轮次类型
            - 导入失败会返回详细错误信息
            
        Excel格式要求:
            - 必填列: 赛事名称、参赛者、轮次、成绩
            - 可选列: 排名
            - 第一行为表头
            - 支持中英文表头
            
        参数:
            - file: Excel文件（.xlsx格式）
            - context_event: 当前筛选的赛事ID（可选，用于验证）
            
        返回:
            - imported: 成功导入的数量
            - errors: 错误信息列表（包含行号和详情）
            
        注意事项:
            - 参赛者必须已有审核通过的报名记录
            - 同一参赛者在同一轮次不能重复导入
            - 赛事名称必须完全匹配
        """
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': '未提供文件'}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file.name.lower().endswith('.xlsx'):
            return Response({'error': '仅支持 .xlsx 文件'}, status=status.HTTP_400_BAD_REQUEST)
        context_event_id = request.data.get('context_event')
        try:
            context_event_id = int(context_event_id) if context_event_id else None
        except (TypeError, ValueError):
            context_event_id = None
        try:
            uploaded_file.seek(0)
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            return Response({'error': '无法读取 Excel 文件', 'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        column_mapping, missing = build_column_mapping(header_row)
        if missing:
            workbook.close()
            return Response({'error': f'缺少必要字段: {", ".join(missing)}'}, status=status.HTTP_400_BAD_REQUEST)
        if sheet.max_row <= 1:
            workbook.close()
            return Response({'error': 'Excel 文件没有数据行'}, status=status.HTTP_400_BAD_REQUEST)
        errors = []
        imported = 0
        event_cache = {}
        try:
            with transaction.atomic():
                for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(cell is not None for cell in row):
                        continue
                    def get_cell(field):
                        idx = column_mapping.get(field)
                        if idx is None or idx >= len(row):
                            return None
                        return row[idx]
                    event_text = safe_str(get_cell('event'))
                    if not event_text:
                        errors.append({'row': row_index, 'detail': '赛事名称为空'})
                        continue
                    cache_key = event_text.lower()
                    event = event_cache.get(cache_key)
                    if event is None:
                        event = Event.objects.filter(title__iexact=event_text).first()
                        if not event:
                            errors.append({'row': row_index, 'detail': f'找不到赛事: {event_text}'})
                            continue
                        event_cache[cache_key] = event
                    if context_event_id and event.id != context_event_id:
                        errors.append({'row': row_index, 'detail': f'行中的赛事与当前筛选赛事不一致 ({event.title})'})
                        continue
                    participant_text = safe_str(get_cell('participant'))
                    candidates = build_candidate_names(participant_text)
                    if not candidates:
                        errors.append({'row': row_index, 'detail': '参赛者信息为空'})
                        continue
                    registration, reg_error = find_registration_for_participant(event, candidates)
                    if reg_error:
                        errors.append({'row': row_index, 'detail': reg_error})
                        continue
                    if not registration:
                        errors.append({'row': row_index, 'detail': '未找到匹配的报名记录'})
                        continue
                    round_value = get_cell('round')
                    normalized_round = normalize_round_type(round_value)
                    if not normalized_round:
                        errors.append({'row': row_index, 'detail': '轮次无法识别'})
                        continue
                    score_text = safe_str(get_cell('score'))
                    if not score_text:
                        errors.append({'row': row_index, 'detail': '成绩为空'})
                        continue
                    rank_value = None
                    rank_cell = get_cell('rank')
                    if rank_cell not in (None, ''):
                        try:
                            rank_value = int(float(rank_cell))
                        except (ValueError, TypeError):
                            errors.append({'row': row_index, 'detail': '排名需要为数字'})
                            continue
                    if Result.objects.filter(event=event, registration=registration, round_type=normalized_round).exists():
                        errors.append({'row': row_index, 'detail': '该参赛者在当前轮次已有成绩'})
                        continue
                    serializer = ResultCreateSerializer(
                        data={
                            'event': event.id,
                            'registration': registration.id,
                            'round_type': normalized_round,
                            'score': score_text,
                            'rank': rank_value
                        },
                        context={'request': request}
                    )
                    if not serializer.is_valid():
                        detail = '; '.join(f"{k}: {v[0]}" for k, v in serializer.errors.items())
                        errors.append({'row': row_index, 'detail': detail})
                        continue
                    serializer.save()
                    imported += 1
        finally:
            workbook.close()
        if imported == 0:
            return Response({'error': '未导入任何成绩', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
        status_code = status.HTTP_201_CREATED if not errors else status.HTTP_200_OK
        return Response({'message': f'已成功导入{imported}条成绩', 'imported': imported, 'errors': errors}, status=status_code)
